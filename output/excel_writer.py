from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.roster import OWNER_COLORS, OUTPUT_COLS, STAT_COLS, PCT_COLS

HEADERS = {
    "Fantasy_Owner": "Owner",    "PLAYER_ID": "Player ID",  "PLAYER": "Player",
    "TEAM": "Team",              "MATCHUP": "Matchup",
    "PTS": "PTS",
    "FGM": "FGM",   "FGA": "FGA",   "FG_PCT": "FG%",
    "FG3M": "3PM",  "FG3A": "3PA",  "FG3_PCT": "3P%",
    "FTM": "FTM",   "FTA": "FTA",   "FT_PCT": "FT%",
    "OREB": "OREB", "DREB": "DREB", "REB": "REB",
    "AST": "AST",   "TO": "TO",     "STL": "STL",  "BLK": "BLK",
}

COL_WIDTHS = {
    "Fantasy_Owner": 10, "PLAYER_ID": 11, "PLAYER": 24,
    "TEAM": 6,           "MATCHUP": 16,
    "PTS": 6,
    "FGM": 6,  "FGA": 6,  "FG_PCT": 7,
    "FG3M": 6, "FG3A": 6, "FG3_PCT": 7,
    "FTM": 6,  "FTA": 6,  "FT_PCT": 7,
    "OREB": 6, "DREB": 6, "REB": 6,
    "AST": 6,  "TO": 6,   "STL": 6,  "BLK": 6,
}

DARK = "16213E"
DIVIDER = "CCCCCC"
CENTER_COLS = STAT_COLS + PCT_COLS + ["TEAM", "Fantasy_Owner", "PLAYER_ID"]


def write_excel(df: pd.DataFrame, target_date: date, out_path: Path):
    """Write game log DataFrame to a formatted .xlsx workbook."""
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "All Players"
    _write_sheet(ws_all, df, target_date, color_by_owner=True)

    for owner in sorted(df["Fantasy_Owner"].unique()):
        ws = wb.create_sheet(title=owner)
        owner_df = df[df["Fantasy_Owner"] == owner].copy()
        _write_sheet(ws, owner_df, target_date, color_by_owner=False,
                     fill_color=OWNER_COLORS.get(owner, "FFFFFF"))

    wb.save(out_path)
    print(f"\n✅  Saved → {out_path}")


def _write_sheet(ws, df: pd.DataFrame, target_date: date,
                 color_by_owner: bool, fill_color: str = "FFFFFF"):
    cols = [c for c in OUTPUT_COLS if c in df.columns]

    _write_title(ws, target_date, len(cols))
    _write_headers(ws, cols)
    last_data_row = _write_data(ws, df, cols, color_by_owner, fill_color)
    _write_totals(ws, cols, last_data_row)
    _set_col_widths(ws, cols)
    ws.freeze_panes = "A3"


def _write_title(ws, target_date: date, num_cols: int):
    title = f"Fantasy Basketball Game Log — {target_date.strftime('%B %d, %Y')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", start_color="1A1A2E")


def _write_headers(ws, cols: list[str]):
    hdr_fill = PatternFill("solid", start_color=DARK)
    for col_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=HEADERS.get(col, col))
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")


def _write_data(ws, df: pd.DataFrame, cols: list[str],
                color_by_owner: bool, fill_color: str) -> int:
    """Write data rows; returns the index of the last row written."""
    prev_owner = None
    row_idx = 3

    for _, row in df[cols].iterrows():
        owner = row.get("Fantasy_Owner", "")

        if color_by_owner:
            if owner != prev_owner and prev_owner is not None:
                for c in range(1, len(cols) + 1):
                    ws.cell(row=row_idx, column=c).fill = PatternFill(
                        "solid", start_color=DIVIDER)
                row_idx += 1
            prev_owner = owner
            row_color = OWNER_COLORS.get(owner, "FFFFFF")
        else:
            row_color = fill_color

        for col_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col])
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=row_color)
            cell.alignment = Alignment(
                horizontal="center" if col in CENTER_COLS else "left"
            )
            if col in PCT_COLS:
                cell.number_format = "0.0%"

        row_idx += 1

    return row_idx - 1  # last row written


def _write_totals(ws, cols: list[str], last_data_row: int):
    if last_data_row < 3:
        return

    totals_row = last_data_row + 2
    label_col = next((i + 1 for i, c in enumerate(cols) if c == "PLAYER"), 1)

    label_cell = ws.cell(row=totals_row, column=label_col, value="TOTALS")
    label_cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    label_cell.fill = PatternFill("solid", start_color=DARK)

    for col_idx, col in enumerate(cols, start=1):
        if col not in STAT_COLS and col not in PCT_COLS:
            continue
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}3:{col_letter}{last_data_row}"

        if col in PCT_COLS:
            # Derive the made/attempted cols that back this percentage
            made_col, att_col = {
                "FG_PCT":  ("FGM",  "FGA"),
                "FG3_PCT": ("FG3M", "FG3A"),
                "FT_PCT":  ("FTM",  "FTA"),
            }[col]
            made_idx = cols.index(made_col) + 1
            att_idx  = cols.index(att_col)  + 1
            made_letter = get_column_letter(made_idx)
            att_letter  = get_column_letter(att_idx)
            formula = (
                f"=IF(SUM({att_letter}3:{att_letter}{last_data_row})=0, 0, "
                f"SUM({made_letter}3:{made_letter}{last_data_row})/"
                f"SUM({att_letter}3:{att_letter}{last_data_row}))"
            )
            number_format = "0.0%"
        else:
            formula = f"=SUM({data_range})"
            number_format = "General"

        cell = ws.cell(row=totals_row, column=col_idx, value=formula)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=DARK)
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = number_format


def _set_col_widths(ws, cols: list[str]):
    for col_idx, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 10)
